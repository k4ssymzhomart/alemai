"""Shared XLSX building blocks (EPIC F2).

House style: bold header row, frozen at A2, explicit column widths, numbers as
numbers (₸ with space thousands via number_format, never strings), dates as
date cells. File naming law: ``qalam_<screen>_<date>.xlsx``.
"""

from __future__ import annotations

import datetime
import io
from dataclasses import dataclass
from typing import Any
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: ₸ integers with space thousands: 6 665 700
TENGE_FMT = "#,##0"
DATE_FMT = "DD.MM.YYYY"
PCT_FMT = "0.0"


@dataclass(frozen=True, slots=True)
class Column:
    header: str
    width: int
    number_format: str | None = None  # e.g. TENGE_FMT / DATE_FMT


def build_xlsx(title: str, columns: list[Column], rows: list[list[Any]]) -> bytes:
    """One-sheet workbook in the house style; values land typed, not stringified."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31] or "Sheet1"

    sheet.append([c.header for c in columns])
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
    sheet.freeze_panes = "A2"

    for i, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = column.width

    for row in rows:
        sheet.append(list(row))
        if any(c.number_format for c in columns):
            for i, column in enumerate(columns, start=1):
                if column.number_format:
                    sheet.cell(row=sheet.max_row, column=i).number_format = (
                        column.number_format
                    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_filename(screen: str, day: datetime.date | None = None) -> str:
    """File naming law: qalam_<screen>_<date>.xlsx."""
    return f"qalam_{screen}_{(day or datetime.date.today()).isoformat()}.xlsx"


def xlsx_response(data: bytes, filename: str) -> StreamingResponse:
    """Wrap XLSX bytes with download headers (RFC 5987 filename*)."""
    return StreamingResponse(
        io.BytesIO(data),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{filename}\"; "
                f"filename*=UTF-8''{quote(filename)}"
            )
        },
    )
