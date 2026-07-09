"""МИС registry import — EPIC F1 (docs/17, ST-7, docs/16 §5 ingest).

Pipeline: parse XLSX/CSV → column mapping via the saved preset «Damumed:
реестр услуг» (auto-map report) → row validation (bad rows → quarantine with
reasons, never silent drops) → match rows to existing claims by natural key →
idempotent upsert (re-upload = no duplicates — the mechanism that makes a live
demo upload safe) → detected-period rules run is triggered by the API layer.

Column shape follows docs/research/schet_reestr_columns.csv +
damumed_export_format.md §4.1 (the реестр услуг reconstruction). Confidence
tags ride along: CONFIRMED = column exists in the official счёт-реестр form
(приказ ҚР ДСМ-86, V2100024058); LIKELY/INFERRED per the research pack.

The natural key is (patient, doctor, service code, date, source, tariff) with
occurrence indexing: the n-th file row carrying a key matches the n-th DB claim
carrying it (both ordered by claim id), so duplicate keys in a real month don't
produce duplicate claims on re-upload.
"""

from __future__ import annotations

import csv
import datetime
import io
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Any

import sqlalchemy as sa
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.claim import Claim
from app.models.enums import ClaimStatus
from app.models.imports import ImportFile, QuarantineRow
from app.models.org import Organization
from app.models.people import Doctor, Patient

PRESET_NAME = "Damumed: реестр услуг"

# ЕПС-style transfer statuses (столбец INFERRED — статус машиночитаемо не
# подтверждён, см. damumed_export_format.md §4.1) ↔ claim status. Bijective,
# so a round-trip export→import is the identity.
STATUS_LABEL_BY_CLAIM: dict[str, str] = {
    ClaimStatus.mis_only.value: "не передан",
    ClaimStatus.submitted.value: "передан",
    ClaimStatus.accepted.value: "принят",
    ClaimStatus.rejected.value: "отклонен",
    ClaimStatus.paid.value: "оплачен",
}
CLAIM_STATUS_BY_LABEL: dict[str, str] = {v: k for k, v in STATUS_LABEL_BY_CLAIM.items()}

# Funding source labels as they appear in выгрузки (ru CONFIRMED in the form
# header; kk variants per shared/glossary.csv).
FUNDING_BY_LABEL: dict[str, str] = {
    "гобмп": "gobmp", "тмккк": "gobmp",
    "осмс": "osms", "мәмс": "osms", "мамс": "osms",
}
FUNDING_LABEL_BY_SOURCE: dict[str, str] = {"gobmp": "ГОБМП", "osms": "ОСМС"}


@dataclass(frozen=True, slots=True)
class PresetColumn:
    """One saved-preset entry: file header → claim field (None = ignored)."""

    header: str
    field: str | None
    confidence: str  # CONFIRMED | LIKELY | INFERRED (docs/research)
    required: bool = False
    note: str | None = None


# The saved preset «Damumed: реестр услуг» (docs/13 §4.7 Интеграции). Headers
# per the официальная форма / research reconstruction; INFERRED columns are our
# best reconstruction of the Damumed выгрузка, not confirmed fields.
DAMUMED_REGISTRY_PRESET: tuple[PresetColumn, ...] = (
    PresetColumn("№ п/п", None, "INFERRED", note="порядковый номер — не хранится"),
    PresetColumn("ИИН пациента", "patient_id", "CONFIRMED", required=True,
                 note="демо: солёный хэш вместо сырого ИИН (персданные не храним)"),
    PresetColumn("ФИО пациента", None, "LIKELY", note="персданные — не храним"),
    PresetColumn("Дата рождения", None, "LIKELY", note="персданные — не храним"),
    PresetColumn("Пол", None, "LIKELY", note="персданные — не храним"),
    PresetColumn("Участок / подразделение", "dept", "INFERRED"),
    PresetColumn("Дата оказания услуги", "date_start", "LIKELY", required=True),
    PresetColumn("Полный код услуги", "service_code", "CONFIRMED", required=True),
    PresetColumn("Наименование услуги", "service_name", "CONFIRMED"),
    PresetColumn("Количество", "qty", "CONFIRMED", required=True),
    PresetColumn("Стоимость услуги, тенге", "tariff", "CONFIRMED", required=True),
    PresetColumn("Сумма, тенге", "amount", "CONFIRMED", required=True),
    PresetColumn("Источник финансирования", "funding_source", "CONFIRMED", required=True),
    PresetColumn("Врач (код)", "doctor_id", "INFERRED", required=True),
    PresetColumn("Диагноз МКБ-10", "icd10", "LIKELY"),
    PresetColumn("Номер направления", "referral_id", "INFERRED"),
    PresetColumn("Статус передачи в ЕПС", "status", "INFERRED"),
)

_PRESET_BY_NORM: dict[str, PresetColumn] = {}


def _norm_header(header: str) -> str:
    return " ".join(str(header).strip().lower().replace("ё", "е").split())


for _col in DAMUMED_REGISTRY_PRESET:
    _PRESET_BY_NORM[_norm_header(_col.header)] = _col


@dataclass(slots=True)
class ColumnMap:
    """Auto-map report entry: which file column matched which system field."""

    column: str
    field: str | None
    status: str  # auto | ignored | unknown
    confidence: str | None = None
    note: str | None = None


@dataclass(slots=True)
class BadRow:
    row_no: int  # 1-based data row number (excluding the header row)
    raw: dict[str, Any]
    errors: list[str]


@dataclass(slots=True)
class RegistryImportReport:
    file_id: uuid.UUID
    filename: str
    preset: str
    mapping: list[ColumnMap]
    period_detected: str | None
    rows_total: int = 0
    rows_ok: int = 0
    matched: int = 0
    updated: int = 0
    new: int = 0
    control_sum: int = 0
    quarantine: list[BadRow] = field(default_factory=list)
    claims_in_period: int = 0


class RegistryParseError(ValueError):
    """The file could not be read as a registry at all (no rows / no header)."""


# ---------------------------------------------------------------------------
# parsing
# ---------------------------------------------------------------------------

def _rows_from_xlsx(data: bytes) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise RegistryParseError("XLSX не содержит листов")
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _rows_from_csv(data: bytes) -> list[list[Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def parse_table(filename: str, data: bytes) -> tuple[list[str], list[list[Any]]]:
    """Read an XLSX/CSV upload into (headers, data rows). Raises RegistryParseError."""
    name = (filename or "").lower()
    if name.endswith((".csv", ".txt")):
        rows = _rows_from_csv(data)
    elif data[:2] == b"PK" or name.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(data)
    else:
        rows = _rows_from_csv(data)
    rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    if not rows:
        raise RegistryParseError("файл пуст — нет ни одной строки")
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    if not any(headers):
        raise RegistryParseError("не найдена строка заголовков")
    return headers, rows[1:]


def map_columns(headers: list[str]) -> tuple[list[ColumnMap], dict[str, int]]:
    """Auto-map file headers via the preset; return the report + field→index."""
    mapping: list[ColumnMap] = []
    index_by_field: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        preset = _PRESET_BY_NORM.get(_norm_header(header))
        if preset is None:
            mapping.append(ColumnMap(column=header, field=None, status="unknown"))
            continue
        if preset.field is None:
            mapping.append(ColumnMap(column=header, field=None, status="ignored",
                                     confidence=preset.confidence, note=preset.note))
            continue
        if preset.field not in index_by_field:
            index_by_field[preset.field] = idx
        mapping.append(ColumnMap(column=header, field=preset.field, status="auto",
                                 confidence=preset.confidence, note=preset.note))
    return mapping, index_by_field


# ---------------------------------------------------------------------------
# row validation
# ---------------------------------------------------------------------------

def _parse_date(value: Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value)) if float(value).is_integer() else int(round(value))
    text = str(value).strip().replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if not text:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


@dataclass(slots=True)
class ParsedRow:
    row_no: int
    raw: dict[str, Any]
    patient_id: str = ""
    doctor_id: str = ""
    dept: str | None = None
    service_code: str = ""
    service_name: str | None = None
    icd10: str | None = None
    date_start: datetime.date | None = None
    qty: int = 1
    tariff: int = 0
    amount: int = 0
    funding_source: str = ""
    referral_id: str | None = None
    status: str = ClaimStatus.submitted.value
    period: str = ""


def _cell(row: list[Any], index_by_field: dict[str, int], field_name: str) -> Any:
    idx = index_by_field.get(field_name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def validate_rows(
    data_rows: list[list[Any]],
    headers: list[str],
    index_by_field: dict[str, int],
    *,
    known_patients: set[str],
    known_doctors: set[str],
    care_type_by_code: dict[str, str],
) -> tuple[list[ParsedRow], list[BadRow]]:
    """Validate every data row; bad rows land in quarantine with all reasons."""
    ok_rows: list[ParsedRow] = []
    bad_rows: list[BadRow] = []
    for i, row in enumerate(data_rows, start=1):
        raw = {headers[j]: row[j] for j in range(min(len(headers), len(row))) if headers[j]}
        raw = {k: (v.isoformat() if isinstance(v, datetime.date | datetime.datetime) else v)
               for k, v in raw.items()}
        errors: list[str] = []
        parsed = ParsedRow(row_no=i, raw=raw)

        parsed.patient_id = str(_cell(row, index_by_field, "patient_id") or "").strip()
        if not parsed.patient_id:
            errors.append("нет ИИН пациента")
        elif parsed.patient_id not in known_patients:
            errors.append("пациент не найден в базе прикрепления (РПН)")

        parsed.doctor_id = str(_cell(row, index_by_field, "doctor_id") or "").strip()
        if not parsed.doctor_id:
            errors.append("нет кода врача")
        elif parsed.doctor_id not in known_doctors:
            errors.append("врач не найден в справочнике")

        parsed.service_code = str(_cell(row, index_by_field, "service_code") or "").strip()
        if not parsed.service_code:
            errors.append("нет кода услуги")
        elif parsed.service_code not in care_type_by_code:
            errors.append(f"код услуги {parsed.service_code} не найден в тарификаторе")

        parsed.date_start = _parse_date(_cell(row, index_by_field, "date_start"))
        if parsed.date_start is None:
            errors.append("дата оказания услуги отсутствует или нечитаема")
        else:
            parsed.period = f"{parsed.date_start.year}-{parsed.date_start.month:02d}"

        qty = _parse_int(_cell(row, index_by_field, "qty"))
        tariff = _parse_int(_cell(row, index_by_field, "tariff"))
        amount = _parse_int(_cell(row, index_by_field, "amount"))
        if qty is None or qty < 1:
            errors.append("количество отсутствует или < 1")
        if tariff is None or tariff < 0:
            errors.append("стоимость услуги отсутствует или отрицательна")
        if amount is None or amount < 0:
            errors.append("сумма отсутствует или отрицательна")
        if None not in (qty, tariff, amount) and amount != qty * tariff:
            errors.append(
                f"сумма не сходится: {amount} ≠ количество {qty} × стоимость {tariff}"
            )
        parsed.qty, parsed.tariff, parsed.amount = qty or 1, tariff or 0, amount or 0

        source_label = str(_cell(row, index_by_field, "funding_source") or "").strip()
        source = FUNDING_BY_LABEL.get(source_label.lower())
        if source is None:
            errors.append(f"источник финансирования не распознан: «{source_label}»")
        else:
            parsed.funding_source = source

        status_label = str(_cell(row, index_by_field, "status") or "").strip().lower()
        if status_label:
            status = CLAIM_STATUS_BY_LABEL.get(status_label)
            if status is None:
                errors.append(f"статус ЕПС не распознан: «{status_label}»")
            else:
                parsed.status = status

        dept = str(_cell(row, index_by_field, "dept") or "").strip()
        parsed.dept = dept or None
        name = str(_cell(row, index_by_field, "service_name") or "").strip()
        parsed.service_name = name or None
        icd10 = str(_cell(row, index_by_field, "icd10") or "").strip()
        parsed.icd10 = icd10 or None
        referral = str(_cell(row, index_by_field, "referral_id") or "").strip()
        parsed.referral_id = referral or None

        if errors:
            bad_rows.append(BadRow(row_no=i, raw=raw, errors=errors))
        else:
            ok_rows.append(parsed)
    return ok_rows, bad_rows


# ---------------------------------------------------------------------------
# matching + idempotent upsert
# ---------------------------------------------------------------------------

NaturalKey = tuple[str, str, str, str, str, int]


def _key_of(patient_id: str, doctor_id: str, service_code: str,
            date_start: datetime.date, funding_source: str, tariff: int) -> NaturalKey:
    return (patient_id, str(doctor_id), service_code, date_start.isoformat(),
            funding_source, tariff)


# Claim fields the registry may legitimately correct on a matched row.
_UPDATABLE = ("service_name", "icd10", "qty", "amount", "referral_id", "dept", "status")


def reference_sets(session: Session) -> tuple[set[str], set[str], dict[str, str]]:
    """Known patients / doctors / service-code→care_type for validation."""
    patients = {p for (p,) in session.execute(sa.select(Patient.id))}
    doctors = {str(d) for (d,) in session.execute(sa.select(Doctor.id))}
    care_type_by_code = {
        code: str(ct)
        for code, ct in session.execute(
            sa.select(Claim.service_code, Claim.care_type).distinct()
        )
    }
    return patients, doctors, care_type_by_code


def upsert_rows(
    session: Session,
    ok_rows: list[ParsedRow],
    care_type_by_code: dict[str, str],
    import_file_id: uuid.UUID,
) -> tuple[int, int, int]:
    """Match ok rows to existing claims by natural key; update diffs, insert rest.

    Returns (matched, updated, new). Idempotent: a re-upload of the same file
    matches every row again and changes nothing.
    """
    periods = sorted({r.period for r in ok_rows})
    if not periods:
        return 0, 0, 0

    existing: dict[NaturalKey, list[Claim]] = defaultdict(list)
    for claim in session.execute(
        sa.select(Claim).where(Claim.period.in_(periods)).order_by(Claim.id)
    ).scalars():
        existing[_key_of(claim.patient_id, str(claim.doctor_id), claim.service_code,
                         claim.date_start, str(claim.funding_source), int(claim.tariff))
                 ].append(claim)

    org_id = session.execute(sa.select(Organization.id).limit(1)).scalar_one()
    taken: Counter[NaturalKey] = Counter()
    matched = updated = new = 0
    for row in ok_rows:
        key = _key_of(row.patient_id, row.doctor_id, row.service_code,
                      row.date_start, row.funding_source, row.tariff)
        candidates = existing.get(key, [])
        occurrence = taken[key]
        taken[key] += 1
        if occurrence < len(candidates):
            claim = candidates[occurrence]
            matched += 1
            changed = False
            for field_name in _UPDATABLE:
                incoming = getattr(row, field_name)
                if field_name in ("service_name", "icd10", "referral_id", "dept") \
                        and incoming is None:
                    continue  # blank cell never wipes existing data
                current = getattr(claim, field_name)
                current = str(current) if field_name == "status" else current
                if incoming != current:
                    setattr(claim, field_name, incoming)
                    changed = True
            if changed:
                updated += 1
        else:
            new += 1
            session.add(Claim(
                org_id=org_id,
                patient_id=row.patient_id,
                doctor_id=uuid.UUID(row.doctor_id),
                dept=row.dept or "—",
                care_type=care_type_by_code[row.service_code],
                funding_source=row.funding_source,
                service_code=row.service_code,
                service_name=row.service_name or row.service_code,
                icd10=row.icd10,
                date_start=row.date_start,
                date_end=row.date_start,
                qty=row.qty,
                tariff=row.tariff,
                amount=row.amount,
                referral_id=row.referral_id,
                status=row.status,
                period=row.period,
                source_file_id=import_file_id,
            ))
    return matched, updated, new


# ---------------------------------------------------------------------------
# orchestration
# ---------------------------------------------------------------------------

def import_registry(session: Session, filename: str, data: bytes) -> RegistryImportReport:
    """Full F1 pipeline. Commits nothing — the caller owns the transaction."""
    headers, data_rows = parse_table(filename, data)
    mapping, index_by_field = map_columns(headers)

    missing = [c.header for c in DAMUMED_REGISTRY_PRESET
               if c.required and c.field not in index_by_field]
    if missing:
        raise RegistryParseError(
            "в файле нет обязательных колонок: " + ", ".join(missing)
        )

    known_patients, known_doctors, care_type_by_code = reference_sets(session)
    ok_rows, bad_rows = validate_rows(
        data_rows, headers, index_by_field,
        known_patients=known_patients,
        known_doctors=known_doctors,
        care_type_by_code=care_type_by_code,
    )

    period_counts = Counter(r.period for r in ok_rows)
    period_detected = period_counts.most_common(1)[0][0] if period_counts else None

    import_file = ImportFile(
        kind="mis",
        filename=filename or "registry.xlsx",
        rows_ok=len(ok_rows),
        rows_quarantined=len(bad_rows),
        control_sum=sum(r.amount for r in ok_rows),
    )
    session.add(import_file)
    session.flush()  # assign import_file.id before claims reference it

    for bad in bad_rows:
        session.add(QuarantineRow(
            import_file_id=import_file.id, row_no=bad.row_no,
            raw=bad.raw, errors=bad.errors,
        ))

    matched, updated, new = upsert_rows(session, ok_rows, care_type_by_code, import_file.id)

    claims_in_period = 0
    if period_detected:
        claims_in_period = session.execute(
            sa.select(sa.func.count()).select_from(Claim)
            .where(Claim.period == period_detected)
        ).scalar_one()

    return RegistryImportReport(
        file_id=import_file.id,
        filename=import_file.filename,
        preset=PRESET_NAME,
        mapping=mapping,
        period_detected=period_detected,
        rows_total=len(data_rows),
        rows_ok=len(ok_rows),
        matched=matched,
        updated=updated,
        new=new,
        control_sum=int(import_file.control_sum or 0),
        quarantine=bad_rows,
        claims_in_period=int(claims_in_period),
    )
