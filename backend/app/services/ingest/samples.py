"""Demo sample files for the Импорт screen (EPIC F1/F3).

Built from the live DB instead of committed binaries so the sample registry is
ALWAYS byte-aligned with the seeded claims — uploading it back reproduces the
golden verdict (46 позиций / 168 600 ₸) with zero drift risk, which is what
makes the live-upload demo beat safe. Shapes follow the «Damumed: реестр
услуг» preset (docs/research/damumed_export_format.md §4.1, INFERRED).

Files:
  registry_2025-11.xlsx — the November registry: exactly the seeded 2025-11
      claims in Damumed-shaped columns.
  registry_broken.xlsx  — 12 rows, 3 deliberately broken (quarantine demo).
  annex_2026.xlsx       — contract annex: current 2026 plan + 2 changed lines
      («Доп. соглашение № 4», preview-only import).
"""

from __future__ import annotations

import datetime
import io

import sqlalchemy as sa
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models.claim import Claim
from app.models.contract import Contract, ContractLine
from app.models.people import Patient
from app.services.ingest.registry import (
    DAMUMED_REGISTRY_PRESET,
    FUNDING_LABEL_BY_SOURCE,
    STATUS_LABEL_BY_CLAIM,
)

_DATE_FMT = "DD.MM.YYYY"
_SEX_LABEL = {"M": "М", "F": "Ж"}

# Annex sample: the 2 changed lines of «Доп. соглашение № 4» (DEMO fixture).
# МРТ gets the volumes beat 3's обращение asked for (storyline 1, пп. 25/26
# п. 19 Правил закупа); стоматология is cut by part of the storyline-2 gap.
ANNEX_CHANGES: dict[tuple[str, str, str], int] = {
    ("kdu", "osms", "МРТ"): 5_700_000,
    ("dent", "osms", ""): -8_700_000,
}

ANNEX_HEADERS = (
    "Вид помощи", "Источник финансирования", "Группа услуг",
    "Годовой план, тенге",
)

CARE_TYPE_LABEL_RU: dict[str, str] = {
    "pmsp": "ПМСП", "kdu": "КДУ", "day_hosp": "Дневной стационар",
    "hosp": "Круглосуточный стационар", "dent": "Стоматология",
    "screening": "Скрининги", "ambulance": "Неотложная помощь",
}
CARE_TYPE_BY_LABEL_RU: dict[str, str] = {v.lower(): k for k, v in CARE_TYPE_LABEL_RU.items()}


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _registry_rows(session: Session, period: str, limit: int | None = None) -> list[list]:
    """Seeded claims of a period as Damumed-shaped row values (preset order)."""
    stmt = (
        sa.select(Claim, Patient.sex, Patient.birth_year)
        .join(Patient, Patient.id == Claim.patient_id)
        .where(Claim.period == period)
        .order_by(Claim.date_start, Claim.id)
    )
    if limit is not None:
        stmt = stmt.limit(limit)
    rows: list[list] = []
    for i, (claim, sex, birth_year) in enumerate(session.execute(stmt), start=1):
        rows.append([
            i,                                            # № п/п
            claim.patient_id,                             # ИИН (демо: хэш)
            "—",                                          # ФИО — не выгружается
            datetime.date(int(birth_year), 1, 1),         # Дата рождения (демо)
            _SEX_LABEL.get(str(sex), str(sex)),           # Пол
            claim.dept,                                   # Участок / подразделение
            claim.date_start,                             # Дата оказания услуги
            claim.service_code,                           # Полный код услуги
            claim.service_name,                           # Наименование услуги
            int(claim.qty),                               # Количество
            int(claim.tariff),                            # Стоимость услуги, тенге
            int(claim.amount),                            # Сумма, тенге
            FUNDING_LABEL_BY_SOURCE[str(claim.funding_source)],
            str(claim.doctor_id),                         # Врач (код)
            claim.icd10,                                  # Диагноз МКБ-10
            claim.referral_id,                            # Номер направления
            STATUS_LABEL_BY_CLAIM[str(claim.status)],     # Статус передачи в ЕПС
        ])
    return rows


def _registry_workbook(rows: list[list], title: str) -> bytes:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=title)
    from openpyxl.cell import WriteOnlyCell  # local: only used here

    header_font = Font(bold=True)
    header_cells = []
    for col in DAMUMED_REGISTRY_PRESET:
        cell = WriteOnlyCell(sheet, value=col.header)
        cell.font = header_font
        header_cells.append(cell)
    sheet.append(header_cells)

    date_cols = {3, 6}  # 0-based: Дата рождения, Дата оказания услуги
    for row in rows:
        cells = []
        for j, value in enumerate(row):
            cell = WriteOnlyCell(sheet, value=value)
            if j in date_cols and isinstance(value, datetime.date):
                cell.number_format = _DATE_FMT
            cells.append(cell)
        sheet.append(cells)
    return _workbook_bytes(workbook)


def registry_sample(session: Session, period: str = "2025-11") -> bytes:
    """The full month registry — exactly the seeded claims of ``period``."""
    return _registry_workbook(_registry_rows(session, period), f"Реестр услуг {period}")


def registry_broken_sample(session: Session, period: str = "2025-11") -> bytes:
    """12 rows, 3 deliberately broken — the quarantine demo file."""
    rows = _registry_rows(session, period, limit=12)
    if len(rows) > 2:
        rows[2][1] = ""                       # row 3: нет ИИН пациента
    if len(rows) > 5:
        rows[5][7] = "S9-XXX"                 # row 6: код не в тарификаторе
        rows[5][8] = "Услуга вне тарификатора"
    if len(rows) > 8:
        rows[8][11] = int(rows[8][11]) + 500  # row 9: сумма ≠ кол-во × стоимость
    return _registry_workbook(rows, f"Реестр услуг {period} (испорчен)")


def annex_sample(session: Session, year: int = 2026) -> bytes:
    """Contract annex: current plan per line-year with 2 changed lines (F3)."""
    plan_rows = session.execute(
        sa.select(
            ContractLine.care_type,
            ContractLine.funding_source,
            ContractLine.service_group,
            sa.func.sum(ContractLine.plan_amount).label("plan_year"),
        )
        .join(Contract, Contract.id == ContractLine.contract_id)
        .where(Contract.year == year)
        .group_by(
            ContractLine.care_type, ContractLine.funding_source,
            ContractLine.service_group,
        )
        .order_by(
            ContractLine.care_type, ContractLine.funding_source,
            ContractLine.service_group,
        )
    ).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Приложение {year} (ДС №4)"
    sheet.append(list(ANNEX_HEADERS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for care_type, source, service_group, plan_year in plan_rows:
        delta = ANNEX_CHANGES.get((str(care_type), str(source), str(service_group or "")), 0)
        sheet.append([
            CARE_TYPE_LABEL_RU.get(str(care_type), str(care_type)),
            FUNDING_LABEL_BY_SOURCE[str(source)],
            str(service_group or ""),
            int(plan_year) + delta,
        ])
    widths = (26, 26, 16, 20)
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(ord("A") + i - 1)].width = width
    return _workbook_bytes(workbook)
