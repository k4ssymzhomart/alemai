#!/usr/bin/env python3
"""Read-back verification of the F2 XLSX exports + F1 samples (openpyxl).

Downloads every export/sample from the live API and asserts: the workbook
opens, the header row is bold, money/qty cells hold numbers (not strings),
and the row counts line up with the API's own JSON answers. AC evidence for
docs/17 EPIC F — run inside the api container:

    docker compose exec -T api python scripts/verify_xlsx_exports.py
"""

from __future__ import annotations

import io
import json
import os
import sys
import urllib.request

from openpyxl import load_workbook

API = os.environ.get("API_BASE", "http://localhost:8000/api/v1")
_fails = 0


def fetch(path: str) -> bytes:
    with urllib.request.urlopen(f"{API}{path}", timeout=120) as r:
        return r.read()


def fetch_json(path: str) -> dict:
    return json.loads(fetch(path))


def check(label: str, ok: bool, detail: str = "") -> None:
    global _fails
    if not ok:
        _fails += 1
    print(f"  [{'PASS' if ok else 'FAIL'}] {label}" + (f": {detail}" if detail else ""))


def sheet_of(data: bytes):
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    return workbook.active


def main() -> int:
    print(f"VERIFY XLSX — {API}")

    # --- prebilling exceptions -------------------------------------------
    data = fetch("/exports/prebilling.xlsx?scope=period:2025-11")
    sheet = sheet_of(data)
    headers = [c.value for c in sheet[1]]
    check("prebilling: opens, has rows", sheet.max_row > 1, f"{sheet.max_row - 1} rows")
    check("prebilling: header row bold", all(c.font.bold for c in sheet[1]))
    check("prebilling: expected columns",
          headers[:3] == ["Правило", "Код ЕКД", "Статус проверки"], str(headers[:3]))
    billed_col = headers.index("Предъявлено, ₸") + 1
    sanction_col = headers.index("Санкция ЕКД, ₸") + 1
    numeric = all(
        isinstance(sheet.cell(row=r, column=billed_col).value, int | float)
        for r in range(2, min(sheet.max_row, 50) + 1)
    )
    check("prebilling: ₸ cells are numbers", numeric)
    blockers = [r for r in range(2, sheet.max_row + 1)
                if str(sheet.cell(row=r, column=3).value or "").startswith("блокер")]
    block_sanction = sum(
        int(sheet.cell(row=r, column=sanction_col).value or 0) for r in blockers
    )
    check("prebilling: 46 blocker rows", len(blockers) == 46, str(len(blockers)))
    check("prebilling: blocker sanction = 6 665 700",
          block_sanction == 6_665_700, f"{block_sanction:,}")

    # --- reconcile bucket 1 ----------------------------------------------
    data = fetch("/exports/reconcile-bucket/1.xlsx")
    sheet = sheet_of(data)
    buckets = fetch_json("/reconcile/buckets")["buckets"]
    bucket1 = next(b for b in buckets if b["bucket_no"] == 1)
    check("bucket1: rows match API", sheet.max_row - 1 == bucket1["rows_count"],
          f"file {sheet.max_row - 1} vs api {bucket1['rows_count']}")
    amount_col = [c.value for c in sheet[1]].index("Сумма, ₸") + 1
    total = sum(int(sheet.cell(row=r, column=amount_col).value or 0)
                for r in range(2, sheet.max_row + 1))
    check("bucket1: Σ amount matches API", total == bucket1["total_amount"],
          f"file {total:,} vs api {bucket1['total_amount']:,}")

    # --- overview ledger ---------------------------------------------------
    data = fetch("/exports/overview.xlsx?year=2026")
    sheet = sheet_of(data)
    overview = fetch_json("/metrics/overview?year=2026")
    check("overview: 13 lines", sheet.max_row - 1 == overview["lines_total"],
          f"{sheet.max_row - 1}")
    headers = [c.value for c in sheet[1]]
    plan_col = headers.index("План на год, ₸") + 1
    plan_total = sum(int(sheet.cell(row=r, column=plan_col).value or 0)
                     for r in range(2, sheet.max_row + 1))
    check("overview: Σ план = plan_amount_year", plan_total == overview["plan_amount_year"],
          f"{plan_total:,}")
    forecast_col = headers.index("Прогноз на год, ₸") + 1
    gap_col = headers.index("Отклонение от плана, ₸") + 1
    gap_total = sum(int(sheet.cell(row=r, column=gap_col).value or 0)
                    for r in range(2, sheet.max_row + 1))
    check("overview: Σ отклонение = forecast_gap", gap_total == overview["forecast_gap"],
          f"{gap_total:,}")
    check("overview: forecast cells are numbers", all(
        sheet.cell(row=r, column=forecast_col).value is None
        or isinstance(sheet.cell(row=r, column=forecast_col).value, int | float)
        for r in range(2, sheet.max_row + 1)
    ))

    # --- samples round-trip shape ------------------------------------------
    data = fetch("/imports/samples/registry_2025-11.xlsx")
    sheet = sheet_of(data)
    check("sample registry: opens", sheet.max_row > 1, f"{sheet.max_row - 1} rows")
    check("sample registry: 17 Damumed columns", sheet.max_column == 17)
    data = fetch("/imports/samples/registry_broken.xlsx")
    sheet = sheet_of(data)
    check("sample broken: 12 rows", sheet.max_row - 1 == 12, f"{sheet.max_row - 1}")
    data = fetch("/imports/samples/annex_2026.xlsx")
    sheet = sheet_of(data)
    check("sample annex: opens, 4 columns", sheet.max_column == 4,
          f"{sheet.max_row - 1} rows")

    print(f"\nVERIFY XLSX: {'ALL PASS' if _fails == 0 else str(_fails) + ' FAIL'}")
    return 1 if _fails else 0


if __name__ == "__main__":
    sys.exit(main())
